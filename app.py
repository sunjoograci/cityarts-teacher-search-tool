"""
CityArts Teacher Finder — Flask web frontend.

Run:
    python app.py
Then open http://localhost:5000
"""
import asyncio
import datetime
import io
import json
import os
import sqlite3
import threading
import time
from pathlib import Path

import requests
from flask import Flask, jsonify, render_template, request, send_file

from src.db import DB_PATH, get_conn, group_teacher_rows, init_db, normalize_person_name
from src.paths import bundle_dir

app = Flask(__name__, template_folder=str(bundle_dir() / "templates"))

init_db()

# ---------------------------------------------------------------------------
# Optional remote scraper service (needed on Vercel, which can't run
# Playwright). If SCRAPER_SERVICE_URL is set, scrape requests are proxied
# there instead of running locally.
# ---------------------------------------------------------------------------
_SCRAPER_SERVICE_URL = os.environ.get("SCRAPER_SERVICE_URL", "").rstrip("/")
_SCRAPE_SHARED_SECRET = os.environ.get("SCRAPE_SHARED_SECRET")

# ---------------------------------------------------------------------------
# Remote ingest: when set, THIS instance's own scrapes push staff results to
# a central server's /api/ingest/school instead of writing to its own local
# database — the "run the scraper on your own machine, residential IP, but
# the team still sees one shared dataset" mode. Absent (the default), this
# app is itself the central server and /api/ingest/school below is what a
# remote-mode instance talks to.
# ---------------------------------------------------------------------------
_REMOTE_INGEST_URL = os.environ.get("REMOTE_INGEST_URL", "").rstrip("/")
_REMOTE_INGEST_SECRET = os.environ.get("REMOTE_INGEST_SECRET") or _SCRAPE_SHARED_SECRET

# ---------------------------------------------------------------------------
# Background scrape job state
# ---------------------------------------------------------------------------

_scrape_lock = threading.Lock()
_scrape_state: dict = {
    "running": False,
    "current": 0,
    "total": 0,
    "current_school": "",
    "last_status": "",
    "started_at": None,
    "finished_at": None,
    "error": None,
    "states": [],
    "rescrape": False,
    "stop_requested": False,
    "saved_this_run": 0,
}


def _run_scrape_thread(states: list[str], rescrape: bool, limit: int | None = None) -> None:
    from src.scraper import run_scraper

    def on_progress(current, total, school_name, status, saved_total):
        _scrape_state["current"] = current
        _scrape_state["total"] = total
        _scrape_state["current_school"] = school_name
        _scrape_state["last_status"] = status
        _scrape_state["saved_this_run"] = saved_total

    def should_stop():
        return _scrape_state["stop_requested"]

    persist_fn = None
    if _REMOTE_INGEST_URL:
        from src.remote_ingest import build_remote_persist_fn
        persist_fn = build_remote_persist_fn(_REMOTE_INGEST_URL, _REMOTE_INGEST_SECRET)

    try:
        asyncio.run(run_scraper(
            states, limit=limit, rescrape_missed=rescrape, on_progress=on_progress,
            should_stop=should_stop, persist_fn=persist_fn,
        ))
    except Exception as exc:
        _scrape_state["error"] = str(exc)
    finally:
        _scrape_state["running"] = False
        _scrape_state["finished_at"] = time.time()


_ON_VERCEL = bool(os.environ.get("VERCEL"))


def _proxy_headers() -> dict:
    return {"X-Scrape-Secret": _SCRAPE_SHARED_SECRET} if _SCRAPE_SHARED_SECRET else {}


@app.route("/api/scrape/start", methods=["POST"])
def api_scrape_start():
    if _SCRAPER_SERVICE_URL:
        resp = requests.post(
            f"{_SCRAPER_SERVICE_URL}/scrape/start",
            json=request.get_json(silent=True) or {},
            headers=_proxy_headers(),
            timeout=15,
        )
        return jsonify(resp.json()), resp.status_code
    if _ON_VERCEL:
        return jsonify({"error": "Scraping isn't configured yet — set SCRAPER_SERVICE_URL to point at a deployed scraper service."}), 403
    with _scrape_lock:
        if _scrape_state["running"]:
            return jsonify({"error": "A scrape is already running."}), 409
        data = request.get_json(silent=True) or {}
        states = [s.upper() for s in data.get("states", ["TX"])]
        rescrape = bool(data.get("rescrape", False))
        limit = data.get("limit")
        limit = int(limit) if limit else None
        _scrape_state.update({
            "running": True,
            "current": 0,
            "total": 0,
            "current_school": "Preparing…",
            "last_status": "",
            "started_at": time.time(),
            "finished_at": None,
            "error": None,
            "states": states,
            "rescrape": rescrape,
            "stop_requested": False,
            "saved_this_run": 0,
        })
    threading.Thread(target=_run_scrape_thread, args=(states, rescrape, limit), daemon=True).start()
    return jsonify({"ok": True})


@app.route("/api/scrape/stop", methods=["POST"])
def api_scrape_stop():
    if _SCRAPER_SERVICE_URL:
        resp = requests.post(
            f"{_SCRAPER_SERVICE_URL}/scrape/stop",
            headers=_proxy_headers(),
            timeout=15,
        )
        return jsonify(resp.json()), resp.status_code
    with _scrape_lock:
        if not _scrape_state["running"]:
            return jsonify({"error": "No scrape is running."}), 409
        _scrape_state["stop_requested"] = True
    return jsonify({"ok": True})


@app.route("/api/scrape/status")
def api_scrape_status():
    if _SCRAPER_SERVICE_URL:
        resp = requests.get(
            f"{_SCRAPER_SERVICE_URL}/scrape/status",
            headers=_proxy_headers(),
            timeout=15,
        )
        return jsonify(resp.json()), resp.status_code
    s = dict(_scrape_state)
    elapsed = None
    eta = None
    started_at_iso = None
    if s["started_at"]:
        import datetime
        elapsed = time.time() - s["started_at"]
        started_at_iso = datetime.datetime.fromtimestamp(
            s["started_at"], tz=datetime.timezone.utc
        ).isoformat()
        if s["current"] > 0 and s["total"] > 0 and s["running"]:
            rate = s["current"] / elapsed
            remaining = s["total"] - s["current"]
            eta = remaining / rate if rate > 0 else None
    s["elapsed"] = elapsed
    s["eta"] = eta
    s["started_at_iso"] = started_at_iso
    # "saved_this_run" (a running total of what persist_fn actually saved,
    # threaded through run_scraper's on_progress) rather than querying this
    # process's own staff table by timestamp: in remote-ingest mode (the
    # desktop app pushing to a shared central server) that table never gets
    # written at all, so the old query-based count always silently read 0
    # regardless of how many teachers were actually found and saved
    # remotely. saved_this_run is accurate in both modes.
    s["new_teachers"] = s.get("saved_this_run", 0)
    return jsonify(s)


def _check_ingest_secret() -> bool:
    if not _SCRAPE_SHARED_SECRET:
        return True
    return request.headers.get("X-Scrape-Secret") == _SCRAPE_SHARED_SECRET


@app.route("/api/ingest/school", methods=["POST"])
def api_ingest_school():
    """Receives one school's scrape results from a remote-mode local
    scraper (see src/remote_ingest.py) and merges them into this server's
    shared database — the endpoint that makes distributed local scraping
    (residential IPs, non-technical teammates) land in one dataset instead
    of N separate SQLite files nobody can merge.

    SECURITY NOTE: if SCRAPE_SHARED_SECRET is not set on this server, this
    endpoint accepts writes from anyone who can reach it (same open-by-
    default posture as /api/scrape/start). That's a materially bigger risk
    here than for scrape/start — an unauthenticated caller could inject
    junk data into the shared database, not just trigger a scrape — so set
    SCRAPE_SHARED_SECRET before pointing any remote scraper at a
    publicly-reachable server.
    """
    if not _check_ingest_secret():
        return jsonify({"error": "unauthorized"}), 401

    # Local import: keeps this module importable without Playwright
    # installed (see run_scraper's import below) — this function only pulls
    # in src.scraper on an actual POST, never at app startup.
    from src.scraper import resolution_method_for_email

    data = request.get_json(silent=True) or {}
    school = data.get("school") or {}
    nces_id = (school.get("nces_id") or "").strip()
    if not nces_id:
        return jsonify({"error": "school.nces_id is required"}), 400

    resolution = data.get("resolution") or {}
    status = data.get("status") or ""
    now = datetime.datetime.now(datetime.timezone.utc).isoformat()

    with get_conn() as conn:
        conn.execute(
            """
            INSERT OR IGNORE INTO schools
                (nces_id, school_name, city, state, district_name, website_url,
                 school_level, is_arts_school)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                nces_id, school.get("school_name") or "", school.get("city"),
                (school.get("state") or "").upper(), school.get("district_name"),
                school.get("website_url"), school.get("school_level"),
                int(bool(school.get("is_arts_school"))),
            ),
        )
        row = conn.execute("SELECT id FROM schools WHERE nces_id=?", (nces_id,)).fetchone()
        school_id = row["id"]

        conn.execute(
            """
            UPDATE schools SET
                entity_type=?, parent_entity=?, parent_entity_type=?,
                resolution_confidence=?, resolution_note=?, domain=?,
                directory_url=?, paths_attempted=?, status=?, needs_human_review=?,
                scraped=1, scrape_status=?
            WHERE id=?
            """,
            (
                resolution.get("entity_type"), resolution.get("parent_entity"),
                resolution.get("parent_entity_type"), resolution.get("resolution_confidence"),
                resolution.get("resolution_note"), resolution.get("domain"),
                data.get("directory_url"), json.dumps(data.get("paths_attempted") or []),
                status, int(bool(resolution.get("needs_human_review"))),
                status, school_id,
            ),
        )

        saved = 0
        for r in data.get("records") or []:
            name = normalize_person_name((r.get("name") or "").strip())
            if not name:
                continue
            email = r.get("email")
            conn.execute(
                """
                INSERT OR IGNORE INTO staff
                    (school_id, teacher_name, title, email, resolution_method,
                     discipline, email_source, email_verified, evidence_url,
                     extraction_strategy, added_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    school_id, name, r.get("title"), email,
                    resolution_method_for_email(email),
                    r.get("discipline"), r.get("email_source"),
                    int(bool(r.get("email_verified"))), r.get("evidence_url"),
                    r.get("extraction_strategy"), now,
                ),
            )
            if conn.execute("SELECT changes()").fetchone()[0]:
                saved += 1

    return jsonify({"ok": True, "school_id": school_id, "saved": saved})


def _get_stats() -> dict:
    if not DB_PATH.exists():
        return {"schools_total": 0, "schools_with_url": 0, "scraped": 0,
                "teachers_total": 0, "teachers_with_email": 0}
    with get_conn() as conn:
        schools_total = conn.execute("SELECT COUNT(*) FROM schools").fetchone()[0]
        schools_with_url = conn.execute(
            "SELECT COUNT(*) FROM schools WHERE website_url IS NOT NULL"
        ).fetchone()[0]
        scraped = conn.execute(
            "SELECT COUNT(*) FROM schools WHERE scraped=1"
        ).fetchone()[0]
        teachers_total = conn.execute("SELECT COUNT(*) FROM staff").fetchone()[0]
        teachers_with_email = conn.execute(
            "SELECT COUNT(*) FROM staff WHERE email IS NOT NULL AND email != ''"
        ).fetchone()[0]
    return {
        "schools_total": schools_total,
        "schools_with_url": schools_with_url,
        "scraped": scraped,
        "teachers_total": teachers_total,
        "teachers_with_email": teachers_with_email,
    }


@app.route("/")
def index():
    stats = _get_stats()
    return render_template("index.html", stats=stats)


@app.route("/api/stats")
def api_stats():
    return jsonify(_get_stats())


@app.route("/api/schools")
def api_schools():
    state = request.args.get("state", "").upper()
    search = request.args.get("q", "").strip()
    only_url = request.args.get("only_url", "0") == "1"
    page = max(1, int(request.args.get("page", 1)))
    per_page = 50

    clauses, params = [], []
    if state:
        clauses.append("state = ?")
        params.append(state)
    if search:
        clauses.append("(school_name LIKE ? OR city LIKE ? OR district_name LIKE ?)")
        params += [f"%{search}%"] * 3
    if only_url:
        clauses.append("website_url IS NOT NULL")

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    with get_conn() as conn:
        total = conn.execute(
            f"SELECT COUNT(*) FROM schools {where}", params
        ).fetchone()[0]
        rows = conn.execute(
            f"""SELECT id, school_name, city, state, district_name,
                       website_url, scraped, scrape_status
                FROM schools {where}
                ORDER BY school_name
                LIMIT ? OFFSET ?""",
            params + [per_page, (page - 1) * per_page],
        ).fetchall()

    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "rows": [dict(r) for r in rows],
    })


@app.route("/api/teachers")
def api_teachers():
    import datetime
    state = request.args.get("state", "").upper()
    search = request.args.get("q", "").strip()
    only_email = request.args.get("only_email", "0") == "1"
    added_since = request.args.get("added_since", "").strip()
    page = max(1, int(request.args.get("page", 1)))
    per_page = 50

    clauses, params = [], []
    if state:
        clauses.append("sc.state = ?")
        params.append(state)
    if search:
        clauses.append(
            "(s.teacher_name LIKE ? OR sc.school_name LIKE ? OR sc.city LIKE ?)"
        )
        params += [f"%{search}%"] * 3
    if only_email:
        clauses.append("s.email IS NOT NULL AND s.email != ''")
    if added_since:
        now = datetime.datetime.now(datetime.timezone.utc)
        if added_since == "today":
            cutoff = now.replace(hour=0, minute=0, second=0, microsecond=0)
        elif added_since == "week":
            cutoff = now - datetime.timedelta(days=7)
        elif added_since == "month":
            cutoff = now - datetime.timedelta(days=30)
        else:
            try:
                cutoff = datetime.datetime.fromisoformat(added_since)
            except ValueError:
                cutoff = None
        if cutoff:
            clauses.append("s.added_at >= ?")
            params.append(cutoff.isoformat())

    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    with get_conn() as conn:
        rows = conn.execute(
            f"""SELECT s.id, s.teacher_name, s.title, s.email, s.resolution_method,
                       sc.school_name, sc.city, sc.state, sc.district_name,
                       sc.website_url, sc.directory_url
                FROM staff s
                JOIN schools sc ON sc.id = s.school_id {where}
                ORDER BY sc.state, sc.school_name, s.teacher_name""",
            params,
        ).fetchall()

    grouped = group_teacher_rows(rows)
    total = len(grouped)
    start = (page - 1) * per_page
    page_rows = grouped[start:start + per_page]

    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "rows": page_rows,
    })


@app.route("/api/art-schools")
def api_art_schools():
    state = request.args.get("state", "").upper()
    search = request.args.get("q", "").strip()
    page = max(1, int(request.args.get("page", 1)))
    per_page = 50

    clauses = ["is_arts_school = 1"]
    params = []
    if state:
        clauses.append("sc.state = ?")
        params.append(state)
    if search:
        clauses.append("(sc.school_name LIKE ? OR sc.city LIKE ? OR sc.district_name LIKE ?)")
        params += [f"%{search}%"] * 3

    where = "WHERE " + " AND ".join(clauses)

    with get_conn() as conn:
        total = conn.execute(
            f"SELECT COUNT(*) FROM schools sc {where}", params
        ).fetchone()[0]
        rows = conn.execute(
            f"""SELECT sc.id, sc.school_name, sc.city, sc.state, sc.district_name,
                       sc.website_url, sc.scraped, sc.scrape_status,
                       COUNT(s.id) as teacher_count
                FROM schools sc
                LEFT JOIN staff s ON s.school_id = sc.id
                {where}
                GROUP BY sc.id
                ORDER BY sc.scraped DESC, sc.school_name
                LIMIT ? OFFSET ?""",
            params + [per_page, (page - 1) * per_page],
        ).fetchall()

    return jsonify({
        "total": total,
        "page": page,
        "per_page": per_page,
        "rows": [dict(r) for r in rows],
    })


@app.route("/api/teachers/<ids>", methods=["DELETE"])
def api_delete_teacher(ids):
    id_list = [int(i) for i in ids.split(",") if i.strip().isdigit()]
    if not id_list:
        return jsonify({"error": "Invalid id"}), 400
    with get_conn() as conn:
        placeholders = ",".join("?" * len(id_list))
        changes = conn.execute(
            f"DELETE FROM staff WHERE id IN ({placeholders})", id_list
        ).rowcount
    if changes:
        return jsonify({"ok": True})
    return jsonify({"error": "Not found"}), 404


@app.route("/api/schools/<int:school_id>", methods=["DELETE"])
def api_delete_school(school_id):
    with get_conn() as conn:
        conn.execute("DELETE FROM staff WHERE school_id = ?", (school_id,))
        changes = conn.execute("DELETE FROM schools WHERE id = ?", (school_id,)).rowcount
    if changes:
        return jsonify({"ok": True})
    return jsonify({"error": "Not found"}), 404


@app.route("/export/teachers.xlsx")
def export_teachers_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    state = request.args.get("state", "").upper()
    only_email = request.args.get("only_email", "0") == "1"

    clauses, params = [], []
    if state:
        clauses.append("sc.state = ?")
        params.append(state)
    if only_email:
        clauses.append("s.email IS NOT NULL AND s.email != ''")
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    with get_conn() as conn:
        rows = conn.execute(
            f"""SELECT s.teacher_name, s.title, s.email, s.resolution_method,
                       sc.school_name, sc.city, sc.state, sc.district_name,
                       sc.website_url, sc.directory_url
                FROM staff s
                JOIN schools sc ON sc.id = s.school_id {where}
                ORDER BY sc.state, sc.school_name, s.teacher_name""",
            params,
        ).fetchall()

    rows = group_teacher_rows(rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Art Teachers"

    headers = [
        "Teacher Name", "Title", "Email", "Contact Method",
        "School", "City", "State", "District", "School Website",
    ]
    header_fill = PatternFill("solid", fgColor="1A3C5E")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    link_font = Font(color="0563C1", underline="single")

    for row in rows:
        email_val = row["email"] or ""
        method = row["resolution_method"] or ""
        is_url = email_val.startswith("http")
        is_form = method in ("send_message_button", "contact_form") or is_url
        # The directory page this teacher was actually found on, not just
        # the school's generic homepage — falls back to the homepage for
        # older rows saved before directory_url was tracked, or a school
        # whose only match was the homepage itself.
        directory_link = row["directory_url"] or row["website_url"] or ""
        ws.append([
            row["teacher_name"],
            row["title"],
            "contact on website" if is_form else email_val,
            method,
            row["school_name"],
            row["city"],
            row["state"],
            row["district_name"],
            directory_link,
        ])
        r = ws.max_row
        if is_url:
            cell = ws.cell(row=r, column=3)
            cell.hyperlink = email_val
            cell.font = link_font
        if directory_link:
            cell = ws.cell(row=r, column=9)
            cell.hyperlink = directory_link
            cell.font = link_font

    col_widths = [22, 24, 28, 16, 32, 16, 6, 28, 36]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"art_teachers{'_' + state if state else ''}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@app.route("/export/schools.xlsx")
def export_schools_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    state = request.args.get("state", "").upper()
    clauses, params = [], []
    if state:
        clauses.append("state = ?")
        params.append(state)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""

    with get_conn() as conn:
        rows = conn.execute(
            f"""SELECT school_name, city, state, district_name,
                       website_url, scraped, scrape_status
                FROM schools {where}
                ORDER BY school_name""",
            params,
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schools"

    headers = ["School", "City", "State", "District", "Website", "Scraped", "Status"]
    header_fill = PatternFill("solid", fgColor="1A3C5E")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([
            row["school_name"],
            row["city"],
            row["state"],
            row["district_name"],
            row["website_url"] or "",
            "Yes" if row["scraped"] else "No",
            row["scrape_status"] or "",
        ])

    col_widths = [32, 16, 6, 28, 40, 8, 24]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"schools{'_' + state if state else ''}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    if not DB_PATH.exists():
        print("No database found. Run: python main.py ingest --states KS")
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=False, host="0.0.0.0", port=port, threaded=True, use_reloader=False)
