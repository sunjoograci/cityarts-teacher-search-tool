"""
Step 1 — School ingestion from NCES CCD public school dataset.

Usage:
    python -m src.ingest --states TX KS [--limit 10]
    python -m src.ingest --file /path/to/ncesdata.xls --states KS
    python -m src.ingest --file /path/to/schools.csv --states TX KS

Accepts CSV, XLSX, and XLS (including the HTML-disguised format exported
by the NCES ELSI Table Generator). When --file is omitted the CCD zip is
downloaded automatically.
"""
import argparse
import csv
import html.parser
import logging
import re
import sqlite3
import zipfile
from pathlib import Path
from typing import Iterator

import requests

from .db import get_conn, init_db, smart_title_case
from .paths import user_data_dir

log = logging.getLogger(__name__)

NCES_CANDIDATE_URLS = [
    "https://nces.ed.gov/sites/default/files/data-asset/ccd-common-core-data/2025/08/2024-25-common-core-data-ccd-preliminary-directory-files/2025046%20Preliminary%20Data%20Release%20CCD%20Nonfiscal_0.zip",
    "https://nces.ed.gov/ccd/Data/zip/ccd_sch_029_1819_w_0a_04082019_csv.zip",
]

DATA_DIR = user_data_dir()
RAW_ZIP = DATA_DIR / "nces_ccd.zip"
RAW_CSV = DATA_DIR / "nces_ccd.csv"

MANUAL_DOWNLOAD_INSTRUCTIONS = """
----------------------------------------------------------------------
MANUAL DOWNLOAD REQUIRED
----------------------------------------------------------------------
The automatic NCES download failed. Please download the school data
manually:

  1. Go to: https://nces.ed.gov/ccd/files.asp
  2. Select:  School Year -> most recent year
              Level -> School
              File Type -> Directory (CSV)
  3. Download the ZIP and extract the CSV.
  4. Save the CSV to:
       {csv_path}
  5. Re-run:  python main.py ingest --states TX KS

Alternatively, use the ELSI Table Generator:
  https://nces.ed.gov/ccd/elsi/tableGenerator.aspx
----------------------------------------------------------------------
"""

# Column-name aliases: maps ELSI and legacy CCD names → canonical names
_COL_ALIASES: dict[str, str] = {
    # ELSI Table Generator column names
    "nces school id": "NCESSCH",
    "school name": "SCH_NAME",
    "district": "LEA_NAME",
    "city": "LCITY",
    "state": "ST",
    "status": "SY_STATUS_TEXT",
    "type": "TYPE",
    # Legacy CCD CSV alternate state column
    "stabr": "ST",
    # Grade level / school level
    "school level": "SCHOOL_LEVEL",
    "school_level": "SCHOOL_LEVEL",
    "level": "SCHOOL_LEVEL",
    "lowest grade offered": "GSLO",
    "highest grade offered": "GSHI",
    "grade span low": "GSLO",
    "grade span high": "GSHI",
    "gslo": "GSLO",
    "gshi": "GSHI",
}

_ART_SCHOOL_RE = re.compile(
    r"\b(arts?|visual arts?|fine arts?|design|creative|performing arts?|"
    r"studio|animation|film|photography|architecture|media arts?|"
    r"digital arts?|graphic arts?|art and)\b",
    re.IGNORECASE,
)


def _grade_to_int(g: str) -> int:
    g = g.strip().upper()
    if g in ("PK", "P", "PR"): return -1
    if g in ("KG", "K"): return 0
    if g == "UN": return -1
    try: return int(g)
    except: return -1


def _is_middle_or_high(row: dict) -> bool:
    """Return True if the school serves any grade 6–12."""
    level = (row.get("SCHOOL_LEVEL") or "").strip()
    if level in ("2", "3"):
        return True
    if level == "1":
        return False
    level_l = level.lower()
    if any(x in level_l for x in ("middle", "high", "secondary", "junior high")):
        return True
    if any(x in level_l for x in ("elementary", "primary", "prekindergarten")):
        return False
    # Grade span fallback
    gshi = (row.get("GSHI") or "").strip()
    if gshi:
        hi = _grade_to_int(gshi)
        if hi >= 6: return True
        if 0 <= hi < 6: return False
    # Name-based fallback — exclude obvious elementary schools
    name = (row.get("SCH_NAME") or "").lower()
    if re.search(r"\b(elementary|primary|elem\b|primaria|pre.?k)\b", name):
        return False
    return True  # unknown → include


def _is_arts_school(school_name: str) -> bool:
    return bool(_ART_SCHOOL_RE.search(school_name))

# HTML entities that indicate missing/N/A data in ELSI exports
_ELSI_NULL = {"†", "–", "‡", "&dagger;", "&ndash;", "&Dagger;", "N", "M", ""}


def _normalize_col(name: str) -> str:
    """Return canonical column name, falling back to the original uppercased."""
    return _COL_ALIASES.get(name.strip().lower(), name.strip())


def _normalize_url(raw: str) -> str | None:
    raw = raw.strip()
    if not raw or raw in _ELSI_NULL:
        return None
    if not raw.startswith("http"):
        raw = "https://" + raw
    return raw


# ---------------------------------------------------------------------------
# File-format parsers — each yields dicts with canonical column names
# ---------------------------------------------------------------------------

class _HTMLTableParser(html.parser.HTMLParser):
    """Minimal SAX-style parser that collects <tr>/<td>/<th> content."""

    def __init__(self) -> None:
        super().__init__()
        self.rows: list[list[str]] = []
        self._current_row: list[str] | None = None
        self._current_cell: list[str] | None = None
        self._in_cell = False

    def handle_starttag(self, tag: str, attrs):
        if tag == "tr":
            self._current_row = []
        elif tag in ("td", "th") and self._current_row is not None:
            self._current_cell = []
            self._in_cell = True

    def handle_endtag(self, tag: str):
        if tag in ("td", "th") and self._current_cell is not None:
            self._current_row.append("".join(self._current_cell).strip())
            self._current_cell = None
            self._in_cell = False
        elif tag == "tr" and self._current_row is not None:
            if any(c.strip() for c in self._current_row):
                self.rows.append(self._current_row)
            self._current_row = None

    def handle_data(self, data: str):
        if self._in_cell and self._current_cell is not None:
            self._current_cell.append(data)

    def handle_entityref(self, name: str):
        if self._in_cell and self._current_cell is not None:
            self._current_cell.append(f"&{name};")


def _rows_from_html_xls(path: Path) -> Iterator[dict]:
    """Parse an HTML-disguised XLS (ELSI export) into row dicts."""
    parser = _HTMLTableParser()
    with open(path, encoding="windows-1252", errors="replace") as fh:
        parser.feed(fh.read())

    # Locate the header row: first row where first cell matches a known NCES header
    header_row_idx = None
    for i, row in enumerate(parser.rows):
        if row and row[0].strip().lower() in ("nces school id", "ncessch", "school id"):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Could not locate header row in NCES HTML export.")

    headers = [_normalize_col(c) for c in parser.rows[header_row_idx]]
    log.info("HTML-XLS headers: %s", headers)

    for row in parser.rows[header_row_idx + 1:]:
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        yield dict(zip(headers, row))


def _rows_from_csv(path: Path) -> Iterator[dict]:
    """Parse a standard NCES CCD CSV."""
    with open(path, encoding="utf-8-sig", errors="replace") as fh:
        reader = csv.DictReader(fh)
        for raw_row in reader:
            yield {_normalize_col(k): v for k, v in raw_row.items()}


def _rows_from_xlsx(path: Path) -> Iterator[dict]:
    """Parse an XLSX file using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for .xlsx files. Install with: pip install openpyxl"
        )
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers: list[str] | None = None
    for row in rows_iter:
        str_row = [str(c).strip() if c is not None else "" for c in row]
        if headers is None:
            # Skip preamble rows; header row has a recognizable first column
            if str_row and str_row[0].lower() in ("nces school id", "ncessch", "school id"):
                headers = [_normalize_col(c) for c in str_row]
            continue
        if len(str_row) < len(headers):
            str_row = str_row + [""] * (len(headers) - len(str_row))
        yield dict(zip(headers, str_row))
    wb.close()


def _is_html(path: Path) -> bool:
    with open(path, "rb") as fh:
        return fh.read(6).lower().startswith(b"<html")


def _iter_rows(path: Path) -> Iterator[dict]:
    """Return an iterator of dicts for any supported file format."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from _rows_from_csv(path)
    elif suffix == ".xlsx":
        yield from _rows_from_xlsx(path)
    elif suffix == ".xls":
        if _is_html(path):
            yield from _rows_from_html_xls(path)
        else:
            # Binary XLS — try xlrd, fall back with a helpful message
            try:
                import xlrd
            except ImportError:
                raise RuntimeError(
                    "xlrd is required for binary .xls files. Install with: pip install xlrd"
                )
            wb = xlrd.open_workbook(str(path))
            ws = wb.sheet_by_index(0)
            headers: list[str] | None = None
            for r in range(ws.nrows):
                str_row = [str(ws.cell_value(r, c)).strip() for c in range(ws.ncols)]
                if headers is None:
                    if str_row and str_row[0].lower() in ("nces school id", "ncessch"):
                        headers = [_normalize_col(c) for c in str_row]
                    continue
                yield dict(zip(headers, str_row))
    else:
        raise ValueError(f"Unsupported file type: {suffix!r}. Use .csv, .xls, or .xlsx")


# ---------------------------------------------------------------------------
# Download logic (used when --file is not provided)
# ---------------------------------------------------------------------------

def download_nces(force: bool = False) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    if RAW_CSV.exists() and not force:
        log.info("NCES CSV already present, skipping download.")
        return

    if not RAW_ZIP.exists() or force:
        downloaded = False
        failures = []
        for url in NCES_CANDIDATE_URLS:
            try:
                log.info("Trying NCES URL: %s", url)
                # requests (not urllib.request) specifically because it
                # bundles its own certifi CA store — a PyInstaller-frozen
                # macOS build doesn't ship the system Python's
                # "Install Certificates.command" trust setup that
                # urllib/ssl's default context relies on, so urlopen()
                # there was failing every HTTPS request with
                # CERTIFICATE_VERIFY_FAILED while silently reporting as a
                # generic "download failed" with no visible reason.
                resp = requests.get(
                    url, headers={"User-Agent": "Mozilla/5.0"}, timeout=30,
                )
                resp.raise_for_status()
                RAW_ZIP.write_bytes(resp.content)
                log.info("Download complete from %s", url)
                downloaded = True
                break
            except Exception as exc:
                log.warning("  Failed (%s): %s", url, exc)
                failures.append(f"{url}: {exc}")

        if not downloaded:
            print(MANUAL_DOWNLOAD_INSTRUCTIONS.format(csv_path=RAW_CSV))
            # RuntimeError, not SystemExit: the desktop app's scrape thread
            # (app.py's _run_scrape_thread) only catches `Exception` to
            # surface it as the scrape's reported error. SystemExit is a
            # BaseException, so it used to skip that except clause
            # entirely and the UI reported the scrape as finished with 0
            # results instead of showing this failure at all.
            raise RuntimeError(
                "Automatic download of the school list failed ("
                + "; ".join(failures) + "). See instructions above/in "
                "MANUAL_DOWNLOAD_INSTRUCTIONS to download manually."
            )

    with zipfile.ZipFile(RAW_ZIP) as zf:
        csv_names = [n for n in zf.namelist() if n.lower().endswith(".csv")]
        if not csv_names:
            raise RuntimeError("No CSV found in NCES zip archive.")
        chosen = _pick_school_csv(zf, csv_names)
        log.info("Extracting %s", chosen)
        with zf.open(chosen) as src, open(RAW_CSV, "wb") as dst:
            dst.write(src.read())
    log.info("Extracted %s", RAW_CSV.name)


def _looks_like_school_csv(zf: zipfile.ZipFile, name: str) -> bool:
    """Peek at just the header row and check it has the columns school
    ingestion actually needs, rather than trusting the filename."""
    with zf.open(name) as fh:
        header_line = fh.readline().decode("utf-8-sig", errors="replace")
    headers = {_normalize_col(c) for c in next(csv.reader([header_line]))}
    return "NCESSCH" in headers and "SCH_NAME" in headers


def _pick_school_csv(zf: zipfile.ZipFile, csv_names: list[str]) -> str:
    """A CCD "Nonfiscal" release zip bundles school-, LEA(district)-, and
    state-level directory files together, and NCES has changed its filename
    convention between releases. Picking by a "_sch_" substring in the
    filename silently chose the wrong file once a release didn't follow
    that convention: it parsed without error but had none of the columns
    ingest_schools() looks for, so every row was skipped and the state
    "ingested" 0 schools with no error anywhere — the scrape then quietly
    reported "0 schools processed" as if it had genuinely found nothing.
    Checking actual header columns is immune to filename changes."""
    for name in csv_names:
        try:
            if _looks_like_school_csv(zf, name):
                return name
        except Exception:
            continue
    # Header sniffing found nothing recognizable (e.g. unexpected
    # encoding) — fall back to the old filename heuristic rather than
    # failing outright.
    school_csvs = [n for n in csv_names if "_sch_" in n.lower()]
    return school_csvs[0] if school_csvs else csv_names[0]


# ---------------------------------------------------------------------------
# Main ingestion
# ---------------------------------------------------------------------------

def ingest_schools(
    states: list[str],
    limit: int | None = None,
    source_file: Path | None = None,
    should_stop=None,
) -> int:
    """`should_stop` is an optional zero-arg callable checked periodically so
    a web-UI Stop request can abort mid-parse — the national CSV is 100k+
    rows and can take minutes on a slow host. Stopping partway is safe
    (INSERT OR IGNORE keyed on nces_id resumes cleanly next run), but the
    caller must then NOT treat the state as fully ingested."""
    if source_file is None:
        download_nces()
        source_file = RAW_CSV

    source_file = Path(source_file)
    if not source_file.exists():
        raise FileNotFoundError(f"Source file not found: {source_file}")

    log.info("Reading schools from %s", source_file)
    init_db()
    states_upper = {s.upper() for s in states}
    inserted = 0

    with get_conn() as conn:
        for row_i, row in enumerate(_iter_rows(source_file)):
            if should_stop is not None and row_i % 2000 == 0 and should_stop():
                log.info("Ingest stopped early by request (%d rows scanned).", row_i)
                break
            state = (row.get("ST") or row.get("STABR") or "").strip().upper()
            if states_upper and state not in states_upper:
                continue
            status = (row.get("SY_STATUS_TEXT") or "").strip()
            if status not in ("Open", "New", "1", "3", ""):
                continue
            if not _is_middle_or_high(row):
                continue
            nces_id = row.get("NCESSCH", "").strip()
            school_name = smart_title_case(row.get("SCH_NAME", "").strip())
            city = smart_title_case(row.get("LCITY", "").strip())
            district_name = smart_title_case(row.get("LEA_NAME", "").strip())
            website_url = _normalize_url(row.get("WEBSITE", ""))
            school_level = (row.get("SCHOOL_LEVEL") or "").strip() or None
            arts = 1 if _is_arts_school(school_name) else 0
            try:
                conn.execute(
                    """
                    INSERT OR IGNORE INTO schools
                        (nces_id, school_name, city, state, district_name, website_url,
                         school_level, is_arts_school)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (nces_id, school_name, city, state, district_name, website_url,
                     school_level, arts),
                )
                if conn.execute("SELECT changes()").fetchone()[0]:
                    inserted += 1
                    if limit and inserted >= limit:
                        break
            except sqlite3.Error as exc:
                log.warning("DB error for %s: %s", school_name, exc)

    log.info("Ingested %d schools for states: %s", inserted, states_upper)
    return inserted


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    parser = argparse.ArgumentParser(description="Ingest NCES school data.")
    parser.add_argument("--states", nargs="+", default=["TX", "KS"], help="State abbreviations")
    parser.add_argument("--limit", type=int, default=None, help="Max schools to ingest (for testing)")
    parser.add_argument("--force", action="store_true", help="Re-download NCES data")
    parser.add_argument(
        "--file",
        type=Path,
        default=None,
        help="Path to a local NCES data file (.csv, .xls, .xlsx). Skips automatic download.",
    )
    args = parser.parse_args()
    n = ingest_schools(args.states, args.limit, source_file=args.file)
    print(f"Done. {n} schools ingested.")


if __name__ == "__main__":
    main()
